"""
Excel Export
════════════
Queries Neo4j for today's leads and writes a formatted .xlsx report.
"""
import logging
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config.settings import OUTPUT_DIR
from utils.db import Neo4jClient

logger = logging.getLogger(__name__)

HEADER_FILL  = PatternFill("solid", fgColor="1B4F72")   # dark blue
HEADER_FONT  = Font(color="FFFFFF", bold=True, size=11)
STRIPE_FILL  = PatternFill("solid", fgColor="D6EAF8")   # light blue
THIN_BORDER  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)


def export_to_excel(since_hours: int = 24) -> Path:
    """
    Pull all leads added in the last `since_hours` hours from Neo4j,
    write them to an Excel file, and return the file path.
    """
    client = Neo4jClient()

    since_dt = datetime.utcnow() - timedelta(hours=since_hours)
    since_iso = since_dt.isoformat()

    rows = client.get_all_leads(since_date=since_iso)
    client.close()

    if not rows:
        logger.warning("No leads found in the last %d hours.", since_hours)
        # Still create an empty report so the email step doesn't fail
        rows = []

    df = pd.DataFrame(rows, columns=[
        "company", "industry", "location", "website",
        "contact_name", "title", "email",
    ])

    # ── File path ─────────────────────────────────────────────────────────────
    date_str = datetime.utcnow().strftime("%Y%m%d")
    filepath = OUTPUT_DIR / f"pharma_leads_{date_str}.xlsx"

    # ── Write to Excel ────────────────────────────────────────────────────────
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Pharma Leads")

    # ── Style the workbook ────────────────────────────────────────────────────
    wb = load_workbook(filepath)
    ws = wb.active

    # Column widths
    col_widths = {
        "A": 30,  # company
        "B": 25,  # industry
        "C": 20,  # location
        "D": 35,  # website
        "E": 25,  # contact_name
        "F": 25,  # title
        "G": 35,  # email
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # Header row styling
    for cell in ws[1]:
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = THIN_BORDER
    ws.row_dimensions[1].height = 20

    # Data rows — alternate stripe + borders
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = STRIPE_FILL if row_idx % 2 == 0 else PatternFill()
        for cell in row:
            cell.fill      = fill
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    wb.save(filepath)
    logger.info("Excel report saved: %s (%d rows)", filepath, len(df))
    return filepath
